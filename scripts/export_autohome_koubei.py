#!/usr/bin/env python3
"""导出汽车之家车型口碑为单 sheet Excel。

功能：
- 抓取指定车型的全部口碑分页
- 逐条进入详情页抓取口碑正文，不抓评论区
- 从详情页正文提取 `评价详情`、`最满意`、`最不满意`
- `综合口碑` 优先按详情页 7 维分数求平均，缺失时回退到列表页值
- 导出为单个 sheet，每条口碑一行

依赖：
- agent-browser 可用
- openpyxl 已安装（仅写 Excel 时需要）
"""

import argparse
import base64
import hashlib
import hmac
import json
import re
import shlex
import subprocess
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from urllib import request as urllib_request
from urllib.parse import urlsplit, urlunsplit

try:
    from tqdm.auto import tqdm
except Exception:  # pragma: no cover
    tqdm = None

HEADERS = [
    "数据类型",
    "用户名",
    "发表日期",
    "口碑标题",
    "综合口碑",
    "车型",
    "行驶里程",
    "电耗",
    "裸车购买价",
    "参考价格",
    "购买时间",
    "探店时间",
    "购买地点",
    "探店地点",
    "评价详情",
    "最满意",
    "最不满意",
    "来源链接",
    "抓取页码",
]

STAGE_WEIGHTS = {
    "detect_pages": 80,
    "collect_reviews": 800,
    "write_excel": 70,
    "write_validation": 50,
}

ALL_REVIEWS_DIMENSION_ID = 0
DETAIL_HEADING_ALIASES = {
    "好评": "最满意",
    "槽点": "最不满意",
}

DETAIL_EVAL_SCRIPT = r"""
(() => {
  const clean = (value) => (value || "").replace(/\s+/g, " ").trim();
  const normalizeHeading = (value) => clean(value).replace(/\s+\d+$/, "").trim();
  const bodyText = clean(document.body.innerText || "");
  const publishedMatch = bodyText.match(/(20\d{2}-\d{2}-\d{2})\s+首次发表/);

  const modelCandidates = Array.from(document.querySelectorAll('a[href*="/spec/"], a[href*="/spec"]'))
    .map((node) => clean(node.textContent))
    .filter((text) => /\d{4}款/.test(text));

  const metaItems = Array.from(document.querySelectorAll(".kb-con li"))
    .map((node) => clean(node.textContent))
    .filter(Boolean);

  const sections = Array.from(document.querySelectorAll(".kb-item"))
    .filter((block) => block.querySelector(".kb-item-msg"))
    .map((block) => {
      const headingNode = block.querySelector("h1, h2, h3, h4");
      const heading = normalizeHeading(headingNode ? headingNode.childNodes[0]?.textContent || headingNode.textContent : "");
      const body = clean(block.querySelector(".kb-item-msg")?.textContent || "");
      const score = clean(block.querySelector(".star-num")?.textContent || "");
      return { heading, body, score };
    })
    .filter((item) => item.body);

  const appendReviews = Array.from(document.querySelectorAll(".conplus"))
    .map((block) => {
      const label = clean(block.previousElementSibling?.textContent || "");
      const body = clean(block.querySelector(".kb-item-msg")?.textContent || "");
      const meta = Array.from(block.querySelectorAll(".car-info li"))
        .map((node) => clean(node.textContent))
        .filter(Boolean);
      return { label, body, meta_items: meta };
    })
    .filter((item) => item.body);

  return JSON.stringify({
    source_link: location.href.split("?")[0].split("#")[0],
    title: clean(document.querySelector("h1.title")?.textContent || ""),
    username: clean(document.querySelector("p.name")?.textContent || ""),
    published_at: publishedMatch ? publishedMatch[1] : "",
    model: modelCandidates[0] || "",
    meta_items: metaItems,
    sections,
    append_reviews: appendReviews
  });
})()
""".strip()


def progress_iter(iterable, *, total=None, desc="", enabled=False):
    if not enabled or tqdm is None:
        return iterable
    return tqdm(iterable, total=total, desc=desc, dynamic_ncols=True, leave=False, file=sys.stderr)


def utc_now_iso():
    return datetime.now(timezone.utc).astimezone().replace(microsecond=0).isoformat()


def write_json_atomic(path, payload):
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = target.with_suffix(target.suffix + ".tmp")
    tmp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_path.replace(target)


def post_json(url, payload, timeout=5):
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib_request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib_request.urlopen(req, timeout=timeout) as resp:
        resp.read()


def make_feishu_signature(secret, timestamp):
    string_to_sign = f"{timestamp}\n{secret}"
    digest = hmac.new(string_to_sign.encode("utf-8"), digestmod=hashlib.sha256).digest()
    return base64.b64encode(digest).decode("utf-8")


def build_feishu_progress_text(payload):
    percent = int(payload.get("percent", 0))
    stage = payload.get("stage") or ""
    stage_progress = payload.get("stage_progress") or {}
    current = int(stage_progress.get("current", payload.get("current", 0)))
    total = int(stage_progress.get("total", payload.get("total", 1)) or 1)
    message = payload.get("message") or ""
    model_name = payload.get("model_name") or ""
    mode = payload.get("mode") or ""
    mode_label = {"autohome_collect": "汽车之家采集"}.get(mode, mode or "任务")
    lines = [f"口碑采集 {percent}%｜{model_name}", f"{mode_label} · {stage} {current}/{total}"]
    if message and message not in {stage, f"{stage} {current}/{total}"}:
        lines.append(f"说明：{message}")
    return "\n".join(lines)


def build_feishu_payload(payload, *, secret=None):
    timestamp = int(time.time())
    feishu_payload = {
        "timestamp": str(timestamp),
        "msg_type": "text",
        "content": {"text": build_feishu_progress_text(payload)},
    }
    if secret:
        feishu_payload["sign"] = make_feishu_signature(secret, timestamp)
    return feishu_payload


class ProgressReporter:
    def __init__(
        self,
        *,
        label,
        model_name,
        mode,
        progress_file=None,
        progress_webhook=None,
        feishu_webhook=None,
        feishu_secret=None,
        total_units=1000,
    ):
        self.label = label
        self.model_name = model_name
        self.mode = mode
        self.progress_file = Path(progress_file).resolve() if progress_file else None
        self.progress_webhook = progress_webhook
        self.feishu_webhook = feishu_webhook
        self.feishu_secret = feishu_secret
        self.total_units = total_units
        self.current_units = 0
        self.stage_base_units = 0
        self.stage_weight_units = 0
        self.stage_total = 1
        self.stage_current = 0
        self.stage_key = "init"
        self.stage_label = "准备"
        self.last_webhook_percent = None
        self.last_webhook_stage = None
        self.disabled_file = False
        self.disabled_webhook = False
        self.disabled_feishu = False

    def _build_payload(self, percent=None, message=""):
        percent = int(percent if percent is not None else round(self.current_units * 100 / self.total_units))
        stage_total = max(int(self.stage_total), 1)
        stage_current = min(int(self.stage_current), stage_total)
        return {
            "label": self.label,
            "mode": self.mode,
            "model_name": self.model_name,
            "stage": self.stage_label,
            "stage_key": self.stage_key,
            "current": int(self.current_units),
            "total": int(self.total_units),
            "percent": percent,
            "message": message,
            "updated_at": utc_now_iso(),
            "overall": {"current": int(self.current_units), "total": int(self.total_units), "percent": percent},
            "stage_progress": {
                "current": stage_current,
                "total": stage_total,
                "percent": int(round(stage_current * 100 / stage_total)),
            },
            "progress_file": str(self.progress_file) if self.progress_file else None,
            "webhook": self.progress_webhook,
        }

    def _emit(self, payload):
        if self.progress_file and not self.disabled_file:
            try:
                write_json_atomic(self.progress_file, payload)
            except Exception as exc:
                self.disabled_file = True
                print(f"WARNING: progress file write failed: {exc}", file=sys.stderr)

        percent = payload["percent"]
        should_emit_webhook = (
            percent != self.last_webhook_percent
            or payload["stage"] != self.last_webhook_stage
            or payload["message"] in {"导出完成", "输入校验失败", "校验失败"}
        )
        if not should_emit_webhook:
            return

        self.last_webhook_percent = percent
        self.last_webhook_stage = payload["stage"]

        if self.progress_webhook and not self.disabled_webhook:
            try:
                post_json(self.progress_webhook, payload)
            except Exception as exc:
                self.disabled_webhook = True
                print(f"WARNING: progress webhook failed: {exc}", file=sys.stderr)
        if self.feishu_webhook and not self.disabled_feishu:
            try:
                post_json(self.feishu_webhook, build_feishu_payload(payload, secret=self.feishu_secret))
            except Exception as exc:
                self.disabled_feishu = True
                print(f"WARNING: feishu webhook failed: {exc}", file=sys.stderr)

    def start_stage(self, stage_key, stage_label, *, weight, total=1, message=""):
        self.stage_key = stage_key
        self.stage_label = stage_label
        self.stage_base_units = self.current_units
        self.stage_weight_units = max(int(weight), 0)
        self.stage_total = max(int(total), 1)
        self.stage_current = 0
        self._emit(self._build_payload(message=message or stage_label))

    def advance(self, current, *, message=""):
        self.stage_current = max(0, min(int(current), self.stage_total))
        fraction = self.stage_current / self.stage_total if self.stage_total else 1.0
        self.current_units = min(self.total_units, int(round(self.stage_base_units + self.stage_weight_units * fraction)))
        self._emit(self._build_payload(message=message or self.stage_label))

    def finish_stage(self, message=""):
        self.stage_current = self.stage_total
        self.current_units = min(self.total_units, int(round(self.stage_base_units + self.stage_weight_units)))
        self._emit(self._build_payload(message=message or f"{self.stage_label}完成"))

    def finalize(self, stage_label="完成", message="导出完成"):
        self.stage_label = stage_label
        self.stage_key = stage_label
        self.stage_current = self.stage_total
        self.current_units = self.total_units
        payload = self._build_payload(percent=100, message=message)
        payload["stage_progress"]["current"] = payload["stage_progress"]["total"]
        payload["stage_progress"]["percent"] = 100
        self._emit(payload)


def run(cmd: str, cwd: Path, timeout: int = 120):
    return subprocess.run(cmd, shell=True, capture_output=True, text=True, cwd=str(cwd), timeout=timeout)


def url_for(series_id: int, page: int, dimensionid: int = ALL_REVIEWS_DIMENSION_ID) -> str:
    if page == 1:
        return f"https://k.autohome.com.cn/{series_id}?dimensionid={dimensionid}&order=0&yearid=0#listcontainer"
    return f"https://k.autohome.com.cn/{series_id}/index_{page}.html?dimensionid={dimensionid}&order=0&yearid=0#listcontainer"


def normalize_detail_url(url: str) -> str:
    parts = urlsplit((url or "").strip())
    return urlunsplit((parts.scheme, parts.netloc, parts.path, "", ""))


def get_page_html(cwd: Path, url: str, session: str, retries: int = 2):
    last_err = ""
    quoted_url = shlex.quote(url)
    for _ in range(retries + 1):
        r1 = run(f"agent-browser --session {session} open {quoted_url} >/dev/null", cwd)
        if r1.returncode != 0:
            last_err = r1.stderr or r1.stdout
            time.sleep(1)
            continue
        r2 = run("agent-browser --session {} eval {}".format(session, shlex.quote("document.documentElement.outerHTML")), cwd)
        if r2.returncode == 0 and r2.stdout.strip():
            return r2.stdout
        last_err = r2.stderr or r2.stdout
        time.sleep(1)
    raise RuntimeError(last_err)


def get_snapshot_any(cwd: Path, url: str, session: str, retries: int = 2):
    last_err = ""
    quoted_url = shlex.quote(url)
    for _ in range(retries + 1):
        r1 = run(f"agent-browser --session {session} open {quoted_url} >/dev/null", cwd)
        if r1.returncode == 0:
            r2 = run(f"agent-browser --session {session} snapshot -c", cwd)
            if r2.returncode == 0 and r2.stdout.strip():
                return r2.stdout.splitlines()
            last_err = r2.stderr or r2.stdout
        else:
            last_err = r1.stderr or r1.stdout
        time.sleep(1)
    raise RuntimeError(last_err)


def run_browser_eval_json(cwd: Path, url: str, session: str, script: str, retries: int = 2):
    last_err = ""
    quoted_url = shlex.quote(url)
    quoted_script = shlex.quote(script)
    for _ in range(retries + 1):
        r1 = run(f"agent-browser --session {session} open {quoted_url} >/dev/null", cwd)
        if r1.returncode != 0:
            last_err = r1.stderr or r1.stdout
            time.sleep(1)
            continue
        r2 = run(f"agent-browser --session {session} eval {quoted_script}", cwd)
        if r2.returncode == 0 and r2.stdout.strip():
            try:
                payload = json.loads(r2.stdout)
                return json.loads(payload) if isinstance(payload, str) else payload
            except json.JSONDecodeError as exc:
                last_err = f"bad json payload: {exc}: {r2.stdout[:400]}"
        else:
            last_err = r2.stderr or r2.stdout
        time.sleep(1)
    raise RuntimeError(last_err)


def detect_max_page(cwd: Path, series_id: int) -> int:
    html = get_page_html(cwd, url_for(series_id, 1), f"koubei_detect_{series_id}")
    candidates = set()
    for pattern in [
        rf"/{series_id}/index_(\d+)\.html\?dimensionid=\d+",
        r"ace-pagination__link[^>]*>(\d+)<",
        r"分页[^\n]{0,200}?共\s*(\d+)\s*页",
        r"共\s*(\d+)\s*页",
        r"尾页[^\n]{0,120}?index_(\d+)\.html",
    ]:
        for match in re.finditer(pattern, html):
            candidates.add(int(match.group(1)))
    candidates = {value for value in candidates if value >= 1}
    return max(candidates) if candidates else 1


def make_empty_row(page: int = 0):
    return {
        "数据类型": "车主购车口碑",
        "用户名": "",
        "发表日期": "",
        "口碑标题": "",
        "综合口碑": "",
        "车型": "",
        "行驶里程": "",
        "电耗": "",
        "裸车购买价": "",
        "参考价格": "",
        "购买时间": "",
        "探店时间": "",
        "购买地点": "",
        "探店地点": "",
        "评价详情": "",
        "最满意": "",
        "最不满意": "",
        "来源链接": "",
        "抓取页码": page,
    }


def norm_user(raw: str) -> str:
    raw = raw.strip()
    if raw.endswith("认证") and " " in raw:
        prefix, suffix = raw.rsplit(" ", 1)
        return f"{prefix}_{suffix}"
    return raw


def extract_cards(lines):
    def is_review_card_start(index: int) -> bool:
        line = lines[index]
        if line.strip() != "- listitem:":
            return False
        if len(line) - len(line.lstrip(" ")) > 4:
            return False
        window = "\n".join(lines[index : index + 8])
        return "https://i.autohome.com.cn/" in window

    stop_index = len(lines)
    for index, line in enumerate(lines):
        if "相关车系口碑推荐" in line:
            stop_index = index
            break

    starts = [index for index in range(stop_index) if is_review_card_start(index)]
    cards = []
    for pos, start in enumerate(starts):
        end = starts[pos + 1] if pos + 1 < len(starts) else stop_index
        block = lines[start:end]
        if any("查看完整口碑" in item for item in block) and any("https://k.autohome.com.cn/detail/view_" in item for item in block):
            cards.append(block)
    return cards


def parse_meta_items_into_row(row: dict, meta_items):
    for item in meta_items:
        value = item.strip()
        if not value:
            continue
        if value.endswith("行驶里程"):
            row["行驶里程"] = value[:-4].strip()
        elif "电耗" in value:
            row["电耗"] = value
        elif value.endswith("裸车购买价"):
            row["裸车购买价"] = value[:-5].strip()
        elif value.endswith("参考价格"):
            row["参考价格"] = value[:-4].strip()
            row["数据类型"] = "试驾探店口碑"
        elif value.endswith("购买时间"):
            row["购买时间"] = value[:-4].strip()
        elif value.endswith("探店时间"):
            row["探店时间"] = value[:-4].strip()
            row["数据类型"] = "试驾探店口碑"
        elif value.endswith("购买地点"):
            row["购买地点"] = value[:-4].strip()
        elif value.endswith("探店地点"):
            row["探店地点"] = value[:-4].strip()
            row["数据类型"] = "试驾探店口碑"


def parse_card_summary(lines, page: int):
    row = make_empty_row(page)
    compact = " ".join(line.strip() for line in lines)

    for index, line in enumerate(lines):
        text = line.strip()

        if text.startswith("- /url: https://i.autohome.com.cn/") and index + 1 < len(lines):
            match = re.match(r'- link "([^"]+)"', lines[index + 1].strip())
            if match and not row["用户名"]:
                row["用户名"] = norm_user(match.group(1))

        if "发表口碑" in text:
            match = re.search(r"(20\d{2}-\d{2}-\d{2})\s+发表口碑", text)
            if match:
                row["发表日期"] = match.group(1)

        if "综合口碑评分" in text:
            match = re.search(r"综合口碑评分\s+([0-9.]+)", text)
            if match:
                row["综合口碑"] = match.group(1)

        match = re.match(r'- link "([^"]+)" \[ref=.*\]:$', text)
        if match and index + 1 < len(lines):
            next_line = lines[index + 1].strip()
            if "https://k.autohome.com.cn/detail/view_" in next_line and not row["口碑标题"]:
                row["口碑标题"] = match.group(1)

        match = re.match(r'- link "(20\d{2}款 [^"]+)"', text)
        if match and not row["车型"]:
            row["车型"] = match.group(1)

        match = re.match(r"- listitem: (.+)", text)
        if match:
            parse_meta_items_into_row(row, [match.group(1).strip()])

        for prefix, data_type, key in [
            ("- text: 满意 ", "车主购车口碑", "最满意"),
            ("- text: 不满意 ", "车主购车口碑", "最不满意"),
            ("- text: 好评 ", "试驾探店口碑", "最满意"),
            ("- text: 槽点 ", "试驾探店口碑", "最不满意"),
        ]:
            if text.startswith(prefix) and not row[key]:
                row[key] = text[len(prefix):].strip()
                row["数据类型"] = data_type

        if text in ("- text: 满意", "- text: 不满意") and not row["最满意"]:
            match = re.search(r"- text: 满意\s+(.*?)\s+- listitem:", compact)
            if match:
                row["最满意"] = match.group(1).strip()

        if text in ("- text: 好评",) and not row["最满意"]:
            match = re.search(r"- text: 好评\s+(.*?)\s+- listitem:", compact)
            if match:
                row["最满意"] = match.group(1).strip()
                row["数据类型"] = "试驾探店口碑"

        if text in ("- text: 不满意",) and not row["最不满意"]:
            match = re.search(r"- text: 不满意\s+(.*?)\s+- listitem:", compact)
            if match:
                row["最不满意"] = match.group(1).strip()

        if text in ("- text: 槽点",) and not row["最不满意"]:
            match = re.search(r"- text: 槽点\s+(.*?)\s+- listitem:", compact)
            if match:
                row["最不满意"] = match.group(1).strip()
                row["数据类型"] = "试驾探店口碑"

        match = re.search(r"https://k\.autohome\.com\.cn/detail/view_[^\s]+", text)
        if match and not row["来源链接"]:
            row["来源链接"] = normalize_detail_url(match.group(0))

    return row


def compose_review_text(sections, append_reviews):
    parts = []
    for section in sections:
        heading = DETAIL_HEADING_ALIASES.get((section.get("heading") or "").strip(), (section.get("heading") or "").strip())
        body = (section.get("body") or "").strip()
        if not body:
            continue
        parts.append(f"{heading}：{body}" if heading else body)

    for append_review in append_reviews:
        label = (append_review.get("label") or "").strip()
        body = (append_review.get("body") or "").strip()
        meta_items = [item.strip() for item in append_review.get("meta_items") or [] if item.strip()]
        if not body:
            continue
        prefix = label or "追加口碑"
        if meta_items:
            prefix = f"{prefix}（{'，'.join(meta_items)}）"
        parts.append(f"{prefix}：{body}")

    return "\n\n".join(parts).strip()


def compute_overall_rating_from_sections(sections) -> str:
    scores = []
    for section in sections:
        raw = (section.get("score") or "").strip()
        if not raw:
            continue
        try:
            scores.append(float(raw))
        except ValueError:
            continue
    if not scores:
        return ""
    return f"{sum(scores) / len(scores):.2f}"


def row_from_detail_payload(payload: dict, page: int):
    row = make_empty_row(page)
    row["口碑标题"] = (payload.get("title") or "").strip()
    row["用户名"] = norm_user((payload.get("username") or "").strip())
    row["发表日期"] = (payload.get("published_at") or "").strip()
    row["车型"] = (payload.get("model") or "").strip()
    row["来源链接"] = normalize_detail_url(payload.get("source_link") or "")

    parse_meta_items_into_row(row, payload.get("meta_items") or [])

    sections = []
    for section in payload.get("sections") or []:
        heading = (section.get("heading") or "").strip()
        body = (section.get("body") or "").strip()
        score = (section.get("score") or "").strip()
        if not body:
            continue
        heading = DETAIL_HEADING_ALIASES.get(heading, heading)
        sections.append({"heading": heading, "body": body, "score": score})

    append_reviews = payload.get("append_reviews") or []
    row["评价详情"] = compose_review_text(sections, append_reviews)
    computed_overall_rating = compute_overall_rating_from_sections(sections)
    row["综合口碑"] = computed_overall_rating or (payload.get("overall_rating") or "").strip()

    for section in sections:
        if section["heading"] == "最满意" and not row["最满意"]:
            row["最满意"] = section["body"]
        if section["heading"] == "最不满意" and not row["最不满意"]:
            row["最不满意"] = section["body"]

    if any([row["参考价格"], row["探店时间"], row["探店地点"]]) or any(
        section["heading"] in {"最满意", "最不满意"} and original in {"好评", "槽点"}
        for section, original in zip(sections, [s.get("heading", "") for s in payload.get("sections") or []], strict=False)
    ):
        row["数据类型"] = "试驾探店口碑"

    return row


def merge_summary_and_detail(summary_row: dict, detail_row: dict):
    merged = make_empty_row(summary_row.get("抓取页码") or detail_row.get("抓取页码") or 0)
    for key in merged:
        if key == "抓取页码":
            merged[key] = detail_row.get(key) or summary_row.get(key) or 0
            continue
        merged[key] = detail_row.get(key) or summary_row.get(key) or ""
    if detail_row.get("数据类型") == "试驾探店口碑" or summary_row.get("数据类型") == "试驾探店口碑":
        merged["数据类型"] = "试驾探店口碑"
    return merged


def validate_row(row: dict) -> bool:
    required = ["用户名", "综合口碑", "车型", "来源链接", "评价详情"]
    return all((row.get(key) or "").strip() for key in required)


def fetch_detail_payload(cwd: Path, url: str):
    session = f"kd_{hashlib.md5(url.encode('utf-8')).hexdigest()[:12]}"
    return run_browser_eval_json(cwd, url, session, DETAIL_EVAL_SCRIPT)


def collect_reviews(cwd: Path, series_id: int, start_page: int, end_page: int, *, reporter=None, show_progress=False):
    rows = []
    bad = []
    page_link_counts = {}
    seen_links = set()

    total_pages = end_page - start_page + 1
    if reporter:
        reporter.start_stage("collect_reviews", "采集全部口碑", weight=STAGE_WEIGHTS["collect_reviews"], total=total_pages, message="开始抓取全部口碑")

    for index, page in enumerate(progress_iter(range(start_page, end_page + 1), total=total_pages, desc="抓取全部口碑", enabled=show_progress), start=1):
        try:
            lines = get_snapshot_any(cwd, url_for(series_id, page), f"koubei_{series_id}_{page}")
        except RuntimeError as exc:
            bad.append({"抓取页码": page, "错误": str(exc), "阶段": "列表页"})
            if reporter:
                reporter.advance(index, message=f"采集全部口碑 {index}/{total_pages}")
            continue

        page_cards = extract_cards(lines)
        page_link_counts[page] = 0
        if not page_cards:
            bad.append({"抓取页码": page, "错误": "未解析到口碑卡片", "阶段": "列表页"})
            if reporter:
                reporter.advance(index, message=f"采集全部口碑 {index}/{total_pages}")
            continue

        for card in page_cards:
            summary_row = parse_card_summary(card, page)
            link = summary_row.get("来源链接", "")
            if not link:
                bad.append({"抓取页码": page, "错误": "卡片缺少详情链接", "阶段": "列表页"})
                continue
            if link in seen_links:
                continue

            seen_links.add(link)
            page_link_counts[page] += 1

            try:
                detail_row = row_from_detail_payload(fetch_detail_payload(cwd, link), page)
            except RuntimeError as exc:
                bad.append({"抓取页码": page, "来源链接": link, "错误": str(exc), "阶段": "详情页"})
                continue

            merged = merge_summary_and_detail(summary_row, detail_row)
            if validate_row(merged):
                rows.append(merged)
            else:
                merged["错误"] = "记录缺少必填字段"
                bad.append(merged)

        if reporter:
            reporter.advance(index, message=f"采集全部口碑 {index}/{total_pages}")

    if reporter:
        reporter.finish_stage("全部口碑采集完成")
    return rows, bad, page_link_counts


def write_xlsx(out_path: Path, rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise SystemExit(f"openpyxl is required to write Excel output: {exc}") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "口碑"
    ws.append(HEADERS)

    for row in rows:
        ws.append([row.get(header, "") for header in HEADERS])

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "S"]:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["O"].width = 110
    ws.column_dimensions["P"].width = 70
    ws.column_dimensions["Q"].width = 70
    ws.column_dimensions["R"].width = 60

    for row in ws.iter_rows(min_row=2):
        for index in [14, 15, 16, 17]:
            row[index].alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(out_path)


def make_reporter(args, *, model_name):
    progress_file = args.progress_file
    if progress_file and not Path(progress_file).is_absolute():
        progress_file = str(Path(progress_file).resolve())
    label = f"koubei-collector:{args.series_id}:{model_name}"
    return ProgressReporter(
        label=label,
        model_name=model_name,
        mode="autohome_collect",
        progress_file=progress_file,
        progress_webhook=args.progress_webhook,
        feishu_webhook=args.feishu_webhook,
        feishu_secret=args.feishu_secret,
    )


def write_validation_report(path: Path, report: dict):
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(description="Export full Autohome koubei to xlsx")
    parser.add_argument("--series-id", type=int, required=True)
    parser.add_argument("--start-page", type=int, default=1)
    parser.add_argument("--end-page", type=int)
    parser.add_argument("--auto-detect-pages", action="store_true", help="Auto detect max page when end page is omitted")
    parser.add_argument("--output", required=True)
    parser.add_argument("--workdir", default=".")
    parser.add_argument("--strict-validate", action="store_true")
    parser.add_argument("--progress", action="store_true", help="Show terminal progress bar")
    parser.add_argument("--progress-file", help="Write progress state to JSON file for polling")
    parser.add_argument("--progress-webhook", help="POST progress state to a generic webhook")
    parser.add_argument("--feishu-webhook", help="POST progress updates to a Feishu incoming webhook")
    parser.add_argument("--feishu-secret", help="Feishu bot secret for signed webhooks")
    args = parser.parse_args()

    cwd = Path(args.workdir).resolve()
    show_progress = args.progress or sys.stderr.isatty()
    use_progress_sink = bool(args.progress_file or args.progress_webhook or args.feishu_webhook)
    reporter = make_reporter(args, model_name=f"series-{args.series_id}") if use_progress_sink else None

    if args.end_page is None:
        if reporter:
            reporter.start_stage("detect_pages", "探测总页数", weight=STAGE_WEIGHTS["detect_pages"], total=1, message="开始自动探测页数")
        if args.auto_detect_pages or args.start_page == 1:
            args.end_page = detect_max_page(cwd, args.series_id)
            print(f"Auto-detected end page: {args.end_page}")
        else:
            raise SystemExit("--end-page is required when start-page is not 1; pass --auto-detect-pages if you want automatic detection")
        if reporter:
            reporter.finish_stage("总页数已探测")

    if args.start_page < 1 or args.end_page < args.start_page:
        raise SystemExit("invalid page range: require start-page >= 1 and end-page >= start-page")

    rows, bad_rows, page_link_counts = collect_reviews(
        cwd,
        args.series_id,
        args.start_page,
        args.end_page,
        reporter=reporter,
        show_progress=show_progress,
    )

    out_path = Path(args.output).resolve()
    if reporter:
        reporter.start_stage("write_excel", "写出 Excel", weight=STAGE_WEIGHTS["write_excel"], total=1, message="开始写出 Excel")
    write_xlsx(out_path, rows)
    if reporter:
        reporter.finish_stage("Excel 已写出")

    report = {
        "ok": len(bad_rows) == 0,
        "mode": "full_reviews_single_sheet",
        "review_total": len(rows),
        "page_link_counts": page_link_counts,
        "raw_parse_anomalies": len(bad_rows),
        "anomalies": bad_rows,
    }
    report_path = out_path.with_suffix(".validation.json")
    if reporter:
        reporter.start_stage("write_validation", "写出校验报告", weight=STAGE_WEIGHTS["write_validation"], total=1, message="开始写出校验报告")
    write_validation_report(report_path, report)
    if reporter:
        reporter.finish_stage("校验报告已写出")
        reporter.finalize(message="导出完成")

    print(f"Wrote: {out_path}")
    print(f"Validation: {report_path}")
    print("Counts:")
    print(f"- 口碑: {len(rows)}")
    print(f"- anomalies: {len(bad_rows)}")
    if report["ok"]:
        print("- validation: OK")
    else:
        print("- validation: FAILED")
        if args.strict_validate:
            raise SystemExit(2)


if __name__ == "__main__":
    main()
